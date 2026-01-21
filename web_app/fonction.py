import csv
import io
import logging as lg
import os
import openpyxl

from collections import Counter
from datetime import date, datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from typing import TypeVar

from .models import (
    Cow,
    Prescription,
    Pharmacie,
    CowUtils,
    PrescriptionUtils,
    PharmacieUtils,
    Traitement,
    UserUtils,
)

basedir = os.path.abspath(os.path.dirname(__file__))
URI = os.path.join(basedir, "txt")

T = TypeVar("T")

def first(lst: list[T]) -> T | None:
    """Renvoie le premier élément d'une liste, ou None si la liste est vide."""
    return lst[0] if lst else None


def last(lst: list[T]) -> T | None:
    """Renvoie le dernier élément d'une liste, ou None si la liste est vide."""
    return lst[-1] if lst else None


def my_strftime(date_obj: date | str) -> str:
    """Convertit une date (objet date ou chaîne de caractères) en une chaîne de
    caractères au format "AAAA-MM-JJ".
    """
    return (date_obj if isinstance(date_obj, date)
                    else datetime.strptime(date_obj, "%Y-%m-%d").date()
            ).strftime('%Y-%m-%d')

def parse_date(date_obj: date | str) -> date:
    """Convertit une chaîne de caractère représentant une date au format
    "AAAA-MM-JJ" en un objet date.
    """
    return (date_obj if isinstance(date_obj, date)
            else datetime.strptime(date_obj, "%Y-%m-%d").date()
            )

def date_to_str(date_obj : date | str) -> str:
    """Convertit  une date ou une chaîne de caractère représentant une date
    au format "AAAA-MM-JJ" en une chaîne de caractère au format "JJ Mois AAA"
    (avec Mois le nom du mois dans la langue du système).
    """
    return (date_obj
            if isinstance(date_obj, date)
            else datetime.strptime(date_obj, "%Y-%m-%d").date()
            ).strftime('%d %B %Y')

def sum_date_to_str(date_obj: date | str, delta_day: int) -> str:
    """Ajoute le nombre de jours passé en argument à la date passée en
    arguments et renvoie le résultat sous la forme d'une chaîne de caractères
    au format "AAAA-MM-JJ".
    """
    return my_strftime(
                parse_date(date_obj) + timedelta(days=delta_day)
            )

def substract_date_to_str(date_obj: date | str, delta_day: int) -> str:
    """Soustrait le nombre de jours passé en argument à la date passée en
    arguments et renvoie le résultat sous la forme d'une chaîne de caractères
    au format "AAAA-MM-JJ".
    """
    return my_strftime(
                parse_date(date_obj) - timedelta(days=delta_day)
            )

def format_bool_fr(value: bool, true_str: str="Oui",
        false_str: str="Non") -> str:
    """Renvoie la transcription des booléens en termes français "Oui" et "Non".
    """
    return true_str if value else false_str

def parse_bool(value: str | None) -> bool | None:
    """Convertit une chaîne de caractères en booléen"""
    if value is None or not value:
        return None
    return value.lower() in {"true", "1", "yes"}

def day_delta(date: date) -> int:
    """Calcule le temps séparant la date d'il y a un an jour pour jour et la
    date fournie en argument.

    Cette fonction détermine le nombre de jours écoulés entre la date d'il y a
    un an jour pour jour et la date fournie en argument. Le résultat est négatif
    si la date fournie précède la date d'il y a un an, positif sinon.

    Arguments:
        * date (date): La date depuis laquelle on cherche à calculer le delta

    Renvoie:
        * int: le nombre de jours entre la date fournie en argument et la date
        un an jour pour jour avant aujourd'hui
    """
    today = datetime.now().date()  # date du jour
    # date d'il y a un ans jour pour jour
    one_year_ago = today - timedelta(days=365)
    return (date - one_year_ago).days

def nb_cares_years(user_id: int, cow_id: int) -> int:
    """Compte le nombre de traitements administrés à une vache au cours de
    l'année passée.

    Cette fonction récupère toutes les entrées de l'historique des traitements
    de la vache associée à l'identifiant fourni en argument et renvoie le nombre
    de ces entrées datant des 365 derniers jours.

    Arguments:
        * user_id (int): Identifiant de l'utilisateur
        * cow_id (int): Identifiant de la vache

    Renvoie:
        * int: le nombre de traitements administrés dans les 365 derniers jours
    """
    cares: list[Traitement] = CowUtils.get_care_by_id(user_id=user_id, cow_id=cow_id)  # type: ignore
    return sum(
        day_delta(parse_date(care["date_traitement"])) >= 0 for care in cares
    )  # sum boolean if True 1 else 0

def nb_cares_years_of_cow(cow: Cow) -> int:
    """Compte le nombre de traitements administrés à une vache au cours de
    l'année passée.

    Cette fonction récupère toutes les entrées de l'historique des traitements
    de la vache représentée par l'objet Cow fourni en argument et renvoie le
    nombre de ces entrées datant des 365 derniers jours.

    Arguments:
        * cow (Cow): L'objet Cow représentant la vache

    Renvoie:
        * int: Le nombre de traitements administrés dans les 365 derniers jours
    """
    cares: list[Traitement] = cow.cow_cares
    return sum(
        day_delta(parse_date(care["date_traitement"])) >= 0 for care in cares
    )  # sum boolean if True 1 else 0

def remaining_care_on_year(cow: Cow) -> int:
    """Calcule le nombre de traitements qu'il est possible d'administrer à la
    vache dans le courant de l'année roulante actuelle.

    Cette fonction calcule le nombre de traitements que la vache représentée par
    l'objet Cow fourni en argument peut encore recevoir dans une période de 365
    jours, sur une base de trois traitements par an.

    Arguments:
        * cow (Cow): objet représentant la vache concernée

    Renvoie:
        * int: Le nombre de traitements que la vache peut encore recevoir dans
        le courant des 365 prochains jours.
    """
    nb_care_year = nb_cares_years_of_cow(cow=cow)
    # traitement restant dans l'année glissante
    return 3 - nb_care_year

def new_available_care(cow: Cow) -> date | None:
    """Determines the next date a cow becomes eligible for a new care treatment.

    This function calculates when a cow can receive its next care treatment based on its care history and the annual treatment limit.

    Args:
        cow (Cow): The cow object whose next available care date is to be determined.

    Returns:
        Optional[date]: The date when the cow is eligible for a new care treatment, or None if no care has been given.
    """

    nb_care_year = nb_cares_years_of_cow(cow=cow)

    if nb_care_year > 0 and len(cow.cow_cares) >= nb_care_year:
        # On prend la date du soin qui correspond à nb_care_year avant la fin de la liste
        care_date = parse_date(cow.cow_cares[-nb_care_year]["date_traitement"])
        return care_date + timedelta(days=365)
    elif len(cow.cow_cares) > 0:
        # Si il y a moins ou autant de soins que nb_care_year, on prend la date du premier soin
        return parse_date(cow.cow_cares[0]["date_traitement"]) + timedelta(days=365)
    else:
        # Pas de soins, donc pas de date dispo
        return None

def get_pharma_list(user_id: int) -> list[str]:
    """Returns a list of all medication names available in the pharmacy.

    This function retrieves the pharmacy list and extracts the medication names from each care item.

    Returns:
        list[str]: A list of medication names.
    """

    return list(UserUtils.get_pharma_list(user_id=user_id))

def get_pharma_len(user_id: int) -> int:
    """Returns the number of medication available in the pharmacy.

    This function counts the total number of unique medications in the pharmacy list.

    Returns:
        int: The number of medication.
    """
    return len(pharma_list) if (pharma_list := get_pharma_list(user_id=user_id)) else 0

def sum_pharmacie_in(user_id: int,year: int) -> dict[str, int]:
    """Sums the quantities of each medication prescribed in a given year.

    This function iterates over all prescriptions for the specified year and accumulates the total quantity for each medication.

    Args:
        year (int): The year to sum medication prescriptions for.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their total prescribed quantities for the year.
    """

    res = {f"{x}": 0 for x in get_pharma_list(user_id=user_id)}
    prescription: Prescription
    for prescription in PrescriptionUtils.get_year_prescription(user_id=user_id, year=year):
        for medic, quantity in prescription.care.items():
            res[medic] += quantity
    return res

def sum_pharmacie_used(user_id: int, year: int) -> dict[str, int]:
    """Sums the quantities of each medication actually used (administered to cows) in a given year.

    This function iterates over all care records for the specified year and accumulates the total quantity used for each medication.

    Args:
        year (int): The year to sum medication usage for.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their total used quantities for the year.
    """
    res = {f"{x}": 0 for x in get_pharma_list(user_id=user_id)}
    cow_care: Traitement
    for cow_care in CowUtils.get_care_on_year(user_id=user_id, year=year):
        for medic, quantity in cow_care["medicaments"].items():
            res[medic] += quantity
    return res

def sum_calf_used(user_id: int, year: int) -> dict[str, int]:
    """Sums the quantities of each medication used for calves in a given year.

    This function iterates over all calf care records for the specified year and accumulates the total quantity used for each medication.

    Args:
        year (int): The year to sum medication usage for calves.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their total used quantities for calves in the year.
    """
    res = {str(x): 0 for x in get_pharma_list(user_id=user_id)}
    cow_care: Traitement
    for cow_care in CowUtils.get_calf_care_on_year(user_id=user_id, year=year):
        lg.info(cow_care)
        for medic, quantity in cow_care["medicaments"].items():
            res[medic] += quantity
    return res

def sum_dlc_left(user_id: int, year: int) -> dict[str, int]:
    """Sums the quantities of each medication removed due to expired shelf life (DLC) in a given year.

    This function iterates over all medication removal records for expired DLC in the specified year and accumulates the total quantity removed for each medication.

    Args:
        year (int): The year to sum medication removals for expired DLC.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their total quantities removed due to expired DLC for the year.
    """
    res = {f"{x}": 0 for x in get_pharma_list(user_id=user_id)}
    cow_care: Prescription
    for cow_care in PrescriptionUtils.get_dlc_left_on_year(user_id=user_id, year=year):
        for medic, quantity in cow_care.care.items():
            res[medic] += quantity
    return res

def sum_pharmacie_left(user_id: int, year: int) -> dict[str, int]:
    """Sums all medications taken out of the pharmacy cabinet in a given year.

    This function adds together the quantities of medications used (administered) and those removed due to expired shelf life (DLC) to give the total quantity of each medication taken out of the pharmacy for the year.

    Args:
        year (int): The year to sum all medication removals from the pharmacy.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their total quantities taken out of the pharmacy for the year.
    """

    return dict(Counter(sum_pharmacie_used(user_id=user_id,year=year)) + Counter(sum_dlc_left(user_id=user_id, year=year)))

def remaining_pharmacie_stock(user_id: int, year: int) -> dict[str, int]:
    """Calculates the remaining stock of each medication in the pharmacy for a given year.

    This function computes the current year's stock by adding the medications prescribed this year and the previous year's remaining stock, then subtracting all medications taken out of the pharmacy this year.

    Args:
        year (int): The year for which to calculate the remaining pharmacy stock.

    Returns:
        dict[str, int]: A dictionary mapping medication names to their remaining quantities for the year.
    """
    #TODO renvoyer le sum de la pharmatie actuel et precedente

    return dict(
        Counter(sum_pharmacie_in(user_id=user_id, year=year))
        + Counter(PharmacieUtils.get_pharmacie_year(user_id=user_id, year=year - 1).remaining_stock)
        - Counter(sum_pharmacie_left(user_id=user_id, year=year))
    )


#TODO def remaining_pharmacie_stock_calcul a fair que sur : traitement + prescription


def get_history_pharmacie(user_id : int) -> list[tuple[date, dict[str, int], str]]:
    """Builds a chronological history of all pharmacy-related events.

    This function combines care and prescription records, labels them, and returns a list sorted by date in descending order.

    Returns:
        list[tuple[date, dict[str:int], str]]: A list of tuples containing the date, medication dictionary, and event type label.
    """

    # Récupère les données
    care_raw : list[tuple[Traitement,int]] = CowUtils.get_all_care(user_id=user_id) or []
    prescription_raw : list[tuple[date, dict[str, int], bool]]= PrescriptionUtils.get_all_prescriptions_cares(user_id=user_id) or []

    care_data = [
        (parse_date(traitement["date_traitement"]), traitement["medicaments"], f"care {cow_id}")
        for traitement, cow_id in care_raw
        ]
    prescription_data = [
        (parse_date(date), medics, "dlc left" if dlc_left else "prescription")
        for date, medics, dlc_left in prescription_raw
    ]

    # Fusionne et trie par date décroissante
    full_history = care_data + prescription_data
    full_history.sort(key=lambda x: x[0], reverse=True)

    return full_history


def update_pharmacie_year(user_id : int, year: int) -> Pharmacie:
    """Updates or creates the pharmacy record for a given year with all relevant medication statistics.

    This function calculates and aggregates medication entries, usages, removals, and remaining stock for the specified year, then updates or creates the corresponding pharmacy record.

    Args:
        year (int): The year for which to update the pharmacy record.

    Returns:
        Pharmacie: The updated or newly created pharmacy record for the year.
    """

    total_enter = sum_pharmacie_in(user_id=user_id, year=year)
    total_used_calf = sum_calf_used(user_id=user_id, year=year)
    total_out_dlc = sum_dlc_left(user_id=user_id, year=year)
    total_used = sum_pharmacie_used(user_id=user_id, year=year)
    total_out = dict(Counter(total_used) + Counter(total_out_dlc))
    remaining_stock = dict(
        Counter(total_enter)
        + Counter(PharmacieUtils.get_pharmacie_year(user_id=user_id, year=year - 1).remaining_stock)
        - Counter(total_out)
    )
    pharmacie = Pharmacie(
        user_id=user_id,
        year=year,
        total_enter=total_enter,
        total_used=total_used,
        total_used_calf=total_used_calf,
        total_out_dlc=total_out_dlc,
        total_out=total_out,
        remaining_stock=remaining_stock,
    )
    return PharmacieUtils.updateOrDefault_pharmacie_year(user_id=user_id, year=year, default=pharmacie)


def pharmacie_to_csv(user_id: int, year: int) -> str:
    """Generates a CSV report of pharmacy medication statistics for a given year.

    This function compiles medication stock, usage, and prescription data into a CSV format, including previous year's stock and per-date prescription details.

    Args:
        year (int): The year for which to generate the pharmacy CSV report.

    Returns:
        str: The generated CSV content as a string.
    """

    pharmacie = update_pharmacie_year(user_id=user_id, year=year)

    # Liste des champs à exporter
    fields = [
        "remaining_stock_last_year",
        "total_enter",
        "total_used",
        "total_used_calf",
        "total_out_dlc",
        "total_out",
        "remaining_stock",
    ]

    # Récupère les données de l'année précédente
    prev_pharmacie = PharmacieUtils.get_pharmacie_year(user_id=user_id, year=year - 1)
    remaining_stock_last_year = getattr(prev_pharmacie, "remaining_stock", {})

    # Obtenir tous les médicaments à partir des données
    all_meds = sorted(get_pharma_list(user_id=user_id))

    output = io.StringIO()
    writer = csv.writer(output)

    # En-tête principale
    writer.writerow(["field"] + all_meds)

    # Ligne spéciale pour l’année précédente
    row = ["remaining_stock_last_year"]
    row.extend(remaining_stock_last_year.get(med, 0) for med in all_meds)
    writer.writerow(row)

    # === AJOUT : lignes des prescriptions par date ===
    # Construire dict : date_str -> med -> qty
    prescriptions_per_date = {
        pres.date.strftime("%d %b %Y"): pres.care # type: ignore
        for pres in PrescriptionUtils.get_year_prescription(user_id=user_id,year=year)
    }

    # Trier les dates
    sorted_dates = sorted(
        prescriptions_per_date.keys(), key=lambda d: datetime.strptime(d, "%d %b %Y")
    )

    # Écrire en CSV avec "prescription DATE" dans la première colonne pour bien identifier
    for date_str in sorted_dates:
        row = [date_str]
        row.extend(prescriptions_per_date[date_str].get(
            med, 0) for med in all_meds)
        writer.writerow(row)

    # === FIN AJOUT ===

    # Autres champs
    for field in fields[1:]:  # on saute 'remaining_stock_last_year' car déjà écrit
        row = [field]
        field_data = getattr(pharmacie, field, {})
        row.extend(field_data.get(med, 0) for med in all_meds)#TODO Verif le get
        writer.writerow(row)

    result = output.getvalue()
    print("CSV généré (pivoté + année précédente + prescriptions par date):\n", result)
    return result


def remaining_care_to_excel(user_id: int) -> bytes:
    """Generates an Excel file summarizing the remaining care treatments for each cow.

    This function creates an Excel spreadsheet listing each cow's ID, the number of remaining treatments, and the next renewal date, with color-coded formatting for easy interpretation.

    Returns:
        bytes: The Excel file content as a bytes object.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traitements Restants" # type: ignore

    headers = ["Numéro Vache",
               "Nb Traitements Restants", "Date Renouvellement"]
    ws.append(headers) # type: ignore

    color_map = {
        3: "00FF00",  # vert
        2: "FFA500",  # orange
        1: "FF0000",  # rouge
        0: "000000",  # noir
    }

    cow: Cow
    for cow in CowUtils.get_all_cows(user_id=user_id):
        cow_id = cow.cow_id
        print("cow_id ",cow.cow_id)
        nb_remaining = remaining_care_on_year(cow)
        print(" "*3, "nb_remaining: ", nb_remaining)
        renewal_date = new_available_care(cow)
        print(" "*3, "renewal_date: ", renewal_date)
        renewal_date_str = renewal_date.strftime(
            "%d %b %Y") if renewal_date else "N/A"
        print(" "*3, "renewal_date_str: ", renewal_date_str)

        ws.append([cow_id, nb_remaining, renewal_date_str]) # type: ignore
        cell = ws.cell(row=ws.max_row, column=2) # type: ignore
        color = color_map.get(nb_remaining, "000000")

        # case coloré
        cell.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid")
        # texte en gras (optionnel)
        cell.font = Font(bold=True)

    # Envoi dans un buffer binaire
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io # type: ignore


def get_all_dry_date(user_id: int) -> dict[int, date]:
    """Retrieves and sorts the dry dates for all cows with valid reproduction records.

    This function collects the 'dry' date for each cow and returns a dictionary sorted by date.

    Returns:
        dict[int, date]: A dictionary mapping cow IDs to their dry dates, sorted by date.
    """
    try:
        dry_dates = {
            cow_id: reproduction["dry"]
            for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=user_id).items()
            if not reproduction.get("dry_status", False)
        }
    except Exception as e:
        lg.error(f"Erreur lors de récupération des dates de tarisement : {e}")
        dry_dates = {}

    return dict(sorted(dry_dates.items(), key=lambda item: item[1]))


def get_all_calving_preparation_date(user_id: int) -> dict[int, date]:
    """Retrieves and sorts the calving preparation dates for all cows with valid reproduction records.

    This function collects the 'calving_preparation' date for each cow and returns a dictionary sorted by date.

    Returns:
        dict[int, date]: A dictionary mapping cow IDs to their calving preparation dates, sorted by date.
    """

    calving_preparation_dates = {
        cow_id: reproduction["calving_preparation"]
        for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=user_id).items()
        if not reproduction["calving_preparation_status"]
    }

    return dict(sorted(calving_preparation_dates.items(), key=lambda item: item[1]))


def get_all_calving_date(user_id: int) -> dict[int, date]:
    """Retrieves and sorts the calving dates for all cows with valid reproduction records.

    This function collects the 'calving_date' for each cow and returns a dictionary sorted by date.

    Returns:
        dict[int, date]: A dictionary mapping cow IDs to their calving dates, sorted by date.
    """

    calving_dates = {
        cow_id: reproduction["calving_date"]
        for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=user_id).items()
    }

    return dict(sorted(calving_dates.items(), key=lambda item: item[1]))


# def rename(pdf: FileStorage, img: FileStorage, article_id: int) -> tuple[str, str]:
#     # Extensions
#     pdf_ext = os.path.splitext(pdf.filename)[1] or ".pdf"
#     img_ext = os.path.splitext(img.filename)[1] or ".jpg"

#     # Noms automatiques
#     pdf_filename = f"blog-{article_id}{pdf_ext}"
#     img_filename = f"blog-{article_id}{img_ext}"

#     # Chemins complets
#     pdf_path = os.path.join('web_Page/static/pdf', pdf_filename)
#     img_path = os.path.join('web_Page/static/images/blog', img_filename)

#     # Sauvegarde
#     pdf.save(pdf_path)
#     img.save(img_path)

#     # Retourner les chemins relatifs depuis "static/"
#     return f"pdf/{pdf_filename}", f"images/blog/{img_filename}"
