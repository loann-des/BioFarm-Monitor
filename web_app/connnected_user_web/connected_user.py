import csv
from enum import Enum
import io
from flask_login import UserMixin
import openpyxl
from openpyxl.styles import Font, PatternFill


from web_app.connnected_user_web.connected_user_dependences_web.CowUtils_user import CowUtilsUser
from web_app.connnected_user_web.connected_user_dependences_web.PrescriptionUtils_user import PrescriptionUtilsUser
from web_app.fonction import date_to_str, day_delta, new_available_care, parse_date, remaining_care_on_year
from ..models.type_dict import (
    Pharma_list_event,
    Prescription_export_format,
    Setting,
    Traitement,
    Traitement_signe
)

from ..models.user import Users, UserUtils
from ..models.cow import CowUtils, Cow
from ..models.prescription import PrescriptionUtils, Prescription
from ..models.pharmacie import PharmacieUtils, Pharmacie
from collections import Counter
from datetime import date, datetime
import logging as lg


class ConnectedUser(UserMixin):
    email: str
    password: str
    setting: Setting
    medic_list : dict
    id: int
    
    cow_utils : "CowUtilsUser"
    
    prescription_utils : "PrescriptionUtilsUser"
        

    def __init__(self, user: Users):
        self.email = user.email
        self.password = user.password
        self.setting = user.setting
        self.id = user.id
        self.medic_list = user.medic_list
        self.cow_ustils = CowUtilsUser(self)
        self.prescription_utils = PrescriptionUtilsUser(self)
        
    def set_user_setting(self, dry_time: int, calving_preparation: int) -> None:
        """Met à jour les paramètres d'élevage de l'utilisateur connecté.

        Cette fonction enregistre les nouvelles durées de tarissement et de
        préparation au vêlage pour l'utilisateur, met à jour les paramètres en
        mémoire, puis recalcule l'ensemble des données de reproduction des
        vaches associées.

        Arguments:
            * dry_time (int): Durée de tarissement en jours
            * calving_preparation (int): Durée de préparation au vêlage en jours
        """
        UserUtils.set_user_setting(
            user_id=self.id, dry_time=dry_time, calving_preparation=calving_preparation
        )
        self.setting["dry_time"] = dry_time
        self.setting["calving_preparation_time"] = calving_preparation
        
        CowUtils.reload_all_reproduction(user_id=self.id, dry_time=dry_time, calving_preparation_time=calving_preparation)


    def add_medic_in_pharma_list(self, medic: str, mesur: str) -> None:
        """Ajoute un médicament à la liste de pharmacie de l'utilisateur connecté.

        Cette fonction délègue à `UserUtils` l'enregistrement d'un nouveau
        médicament et de son unité de mesure dans la configuration de
        l'utilisateur.

        Arguments:
            * medic (str): Nom du médicament à ajouter
            * mesur (str): Unité de mesure associée au médicament
        """
        UserUtils.add_medic_in_pharma_list(
            self.id, medic=medic, mesur=mesur)
    
    def nb_cares_years(self, cow_id: int) -> int:
        """Compte le nombre de traitements administrés à une vache au cours de
        l'année passée.

        Cette fonction récupère toutes les entrées de l'historique des traitements
        de la vache associée à l'identifiant fourni en argument et renvoie le nombre
        de ces entrées datant des 365 derniers jours.

        Arguments:
            * user_id (int): Identifiant de l'utilisateur
            * cow_id (int): Identifiant de la vache

        Renvoie:
            * int: le nombre de traitements administrés dans les 365 derniers jours
        """
        cares: list[Traitement] | None = CowUtils.get_care_by_id(user_id=self.id, cow_id=cow_id)
        return sum(
            day_delta(parse_date(care["date_traitement"])) >= 0 for care in cares
        ) if cares else 0 # sum boolean if True 1 else 0

    def get_pharma_list(self) -> list[str]:
        """Returns a list of all medication names available in the pharmacy.

        This function retrieves the pharmacy list and extracts the medication names from each care item.

        Returns:
            list[str]: A list of medication names.
        """

        return list(UserUtils.get_pharma_list(user_id=self.id))

    def get_pharma_len(self) -> int:
        """Returns the number of medication available in the pharmacy.

        This function counts the total number of unique medications in the pharmacy list.

        Returns:
            int: The number of medication.
        """
        return len(self.get_pharma_list())

    def sum_pharmacie_in(self, year: int) -> dict[str, int]:
        """Sums the quantities of each medication prescribed in a given year.

        This function iterates over all prescriptions for the specified year and accumulates the total quantity for each medication.

        Args:
            year (int): The year to sum medication prescriptions for.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their total prescribed quantities for the year.
        """

        res = {f"{x}": 0 for x in self.get_pharma_list()}
        prescription: Prescription
        for prescription in PrescriptionUtils.get_year_prescription(user_id=self.id, year=year):
            for medic, quantity in prescription.care.items():
                res[medic] += quantity
        return res

    def sum_pharmacie_used(self, year: int) -> dict[str, int]:
        """Sums the quantities of each medication actually used (administered to cows) in a given year.

        This function iterates over all care records for the specified year and accumulates the total quantity used for each medication.

        Args:
            year (int): The year to sum medication usage for.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their total used quantities for the year.
        """
        res = {f"{x}": 0 for x in self.get_pharma_list()}
        cow_care: Traitement
        for cow_care in CowUtils.get_care_on_year(user_id=self.id, year=year):
            for medic, quantity in cow_care["medicaments"].items():
                res[medic] += quantity
        return res

    def sum_calf_used(self, year: int) -> dict[str, int]:
        """Sums the quantities of each medication used for calves in a given year.

        This function iterates over all calf care records for the specified year and accumulates the total quantity used for each medication.

        Args:
            year (int): The year to sum medication usage for calves.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their total used quantities for calves in the year.
        """
        res = {str(x): 0 for x in self.get_pharma_list()}
        cow_care: Traitement
        for cow_care in CowUtils.get_calf_care_on_year(user_id=self.id, year=year):
            lg.info(cow_care)
            for medic, quantity in cow_care["medicaments"].items():
                res[medic] += quantity
        return res

    def sum_dlc_left(self, year: int) -> dict[str, int]:
        """Sums the quantities of each medication removed due to expired shelf life (DLC) in a given year.

        This function iterates over all medication removal records for expired DLC in the specified year and accumulates the total quantity removed for each medication.

        Args:
            year (int): The year to sum medication removals for expired DLC.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their total quantities removed due to expired DLC for the year.
        """
        res = {f"{x}": 0 for x in self.get_pharma_list()}
        cow_care: Prescription
        for cow_care in PrescriptionUtils.get_dlc_left_on_year(user_id=self.id, year=year):
            for medic, quantity in cow_care.care.items():
                res[medic] += quantity
        return res

    def sum_pharmacie_left(self, year: int) -> dict[str, int]:
        """Sums all medications taken out of the pharmacy cabinet in a given year.

        This function adds together the quantities of medications used (administered) and those removed due to expired shelf life (DLC) to give the total quantity of each medication taken out of the pharmacy for the year.

        Args:
            year (int): The year to sum all medication removals from the pharmacy.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their total quantities taken out of the pharmacy for the year.
        """

        return dict(Counter(self.sum_pharmacie_used(year=year)) + Counter(self.sum_dlc_left(year=year)))

    def remaining_pharmacie_stock(self, year: int) -> dict[str, int]:
        """Calculates the remaining stock of each medication in the pharmacy for a given year.

        This function computes the current year's stock by adding the medications prescribed this year and the previous year's remaining stock, then subtracting all medications taken out of the pharmacy this year.

        Args:
            year (int): The year for which to calculate the remaining pharmacy stock.

        Returns:
            dict[str, int]: A dictionary mapping medication names to their remaining quantities for the year.
        """
        return dict(
            Counter(self.sum_pharmacie_in(year=year))
            + Counter(PharmacieUtils.get_pharmacie_year(user_id=self.id,
                      year=year - 1).remaining_stock)
            - Counter(self.sum_pharmacie_left(year=year))
        )

    def get_history_pharmacie(self) -> list[Pharma_list_event]:
        """Builds a chronological history of all pharmacy-related events.

        This function combines care and prescription records, labels them, and returns a list sorted by date in descending order.

        Returns:
            list[tuple[date, dict[str:int], str]]: A list of tuples containing the date, medication dictionary, and event type label.
        """

        # Récupère les données
        cares_raw: list[Traitement_signe] = CowUtils.get_all_care(
            user_id=self.id) or []
        prescriptions_raw: list[Prescription_export_format] = PrescriptionUtils.get_all_prescriptions_cares(
            user_id=self.id) or []

        care_data: list[Pharma_list_event] = [
            Pharma_list_event(date=care_raw["traitement"]["date_traitement"],
                              medicaments=care_raw["traitement"]["medicaments"],
                              event_type=f"care {care_raw['cow_id']}")
            for care_raw in cares_raw
        ]
        prescription_data: list[Pharma_list_event] = [
            Pharma_list_event(date=prescription["date_prescription"],
                              medicaments=prescription["prescription"],
                              event_type="dlc left" if prescription["dlc_left"] else "prescription")

            for prescription in prescriptions_raw
        ]

        # Fusionne et trie par date décroissante
        full_history = care_data + prescription_data
        full_history.sort(key=lambda x: parse_date(x["date"]), reverse=True)

        return full_history

    def update_pharmacie_year(self, year: int) -> Pharmacie:
        """Updates or creates the pharmacy record for a given year with all relevant medication statistics.

        This function calculates and aggregates medication entries, usages, removals, and remaining stock for the specified year, then updates or creates the corresponding pharmacy record.

        Args:
            year (int): The year for which to update the pharmacy record.

        Returns:
            Pharmacie: The updated or newly created pharmacy record for the year.
        """

        total_enter = self.sum_pharmacie_in(year=year)
        total_used_calf = self.sum_calf_used(year=year)
        total_out_dlc = self.sum_dlc_left(year=year)
        total_used = self.sum_pharmacie_used(year=year)
        total_out = dict(Counter(total_used) + Counter(total_out_dlc))
        remaining_stock = dict(
            Counter(total_enter)
            + Counter(PharmacieUtils.get_pharmacie_year(user_id=self.id,
                      year=year - 1).remaining_stock)
            - Counter(total_out)
        )
        pharmacie = Pharmacie(
            user_id=self.id,
            year=year,
            total_enter=total_enter,
            total_used=total_used,
            total_used_calf=total_used_calf,
            total_out_dlc=total_out_dlc,
            total_out=total_out,
            remaining_stock=remaining_stock,
        )
        return PharmacieUtils.updateOrDefault_pharmacie_year(user_id=self.id, default=pharmacie)

    def pharmacie_to_csv(self, year: int) -> str:
        """Generates a CSV report of pharmacy medication statistics for a given year.

        This function compiles medication stock, usage, and prescription data into a CSV format, including previous year's stock and per-date prescription details.

        Args:
            year (int): The year for which to generate the pharmacy CSV report.

        Returns:
            str: The generated CSV content as a string.
        """

        pharmacie = self.update_pharmacie_year(year=year)

        # Liste des champs à exporter
        fields = [
            "remaining_stock_last_year",
            "total_enter",
            "total_used",
            "total_used_calf",
            "total_out_dlc",
            "total_out",
            "remaining_stock",
        ]

        # Récupère les données de l'année précédente
        prev_pharmacie = PharmacieUtils.get_pharmacie_year(
            user_id=self.id, year=year - 1)
        remaining_stock_last_year = getattr(
            prev_pharmacie, "remaining_stock", {})

        # Obtenir tous les médicaments à partir des données
        all_meds = sorted(self.get_pharma_list())
        output = io.StringIO()
        writer = csv.writer(output)

        # En-tête principale
        writer.writerow(["field"] + all_meds)

       # Ligne spéciale pour l’année précédente
        row = ["remaining_stock_last_year"]
        row.extend(remaining_stock_last_year.get(med, 0) for med in all_meds)
        writer.writerow(row)

        # === AJOUT : lignes des prescriptions par date ===
        # Construire dict : date -> med -> qty
        prescriptions_per_date: dict[date, dict[str, int]] = {
            prescription.date: prescription.care
            for prescription in PrescriptionUtils.get_year_prescription(user_id=self.id, year=year)
        }

        # Trier les dates
        sorted_dates = sorted(
            prescriptions_per_date.keys(), key=lambda d: d
        )

        # Écrire en CSV avec "prescription DATE" dans la première colonne pour bien identifier
        date_row: date
        for date_row in sorted_dates:
            row = [date_to_str(date_row)]
            row.extend(str(prescriptions_per_date[date_row].get(
                med, 0)) for med in all_meds)  # TODO Verif le get
            writer.writerow(row)

        # === FIN AJOUT ===

        # Autres champs
        for field in fields[1:]:  # on saute 'remaining_stock_last_year' car déjà écrit
            row = [field]
            field_data = getattr(pharmacie, field, {})
            row.extend(field_data.get(med, 0)
                       for med in all_meds)  # TODO Verif le get
            writer.writerow(row)

        result = output.getvalue()
        print("CSV généré (pivoté + année précédente + prescriptions par date):\n", result)
        return result

    def remaining_care_to_excel(self) -> bytes:
        """Generates an Excel file summarizing the remaining care treatments for each cow.

        This function creates an Excel spreadsheet listing each cow's ID, the number of remaining treatments, and the next renewal date, with color-coded formatting for easy interpretation.

        Returns:
            bytes: The Excel file content as a bytes object.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Traitements Restants"  # type: ignore

        headers = ["Numéro Vache",
                   "Nb Traitements Restants", "Date Renouvellement"]
        ws.append(headers)  # type: ignore

        color_map = {
            3: "00FF00",  # vert
            2: "FFA500",  # orange
            1: "FF0000",  # rouge
            0: "000000",  # noir
        }

        cow: Cow
        for cow in CowUtils.get_all_cows(user_id=self.id):
            cow_id = cow.cow_id
            print("cow_id ", cow.cow_id)
            nb_remaining = remaining_care_on_year(cow)
            print(" "*3, "nb_remaining: ", nb_remaining)
            renewal_date = new_available_care(cow)
            print(" "*3, "renewal_date: ", renewal_date)
            renewal_date_str = renewal_date.strftime(
                "%d %b %Y") if renewal_date else "N/A"
            print(" "*3, "renewal_date_str: ", renewal_date_str)

            ws.append([cow_id, nb_remaining, renewal_date_str])  # type: ignore
            cell = ws.cell(row=ws.max_row, column=2)  # type: ignore
            color = color_map.get(nb_remaining, "000000")

            # case coloré
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
            # texte en gras (optionnel)
            cell.font = Font(bold=True)

        # Envoi dans un buffer binaire
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        return excel_io  # type: ignore

    def get_all_dry_date(self) -> dict[int, date]:
        """Retrieves and sorts the dry dates for all cows with valid reproduction records.

        This function collects the 'dry' date for each cow and returns a dictionary sorted by date.

        Returns:
            dict[int, date]: A dictionary mapping cow IDs to their dry dates, sorted by date.
        """
        try:
            dry_dates : dict[int, date] = {
                cow_id: parse_date(reproduction["dry"])
                for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=self.id).items()
                if not reproduction["dry_status"] and reproduction["dry"] is not None
            }
        except Exception as e:
            lg.error(
                f"Erreur lors de récupération des dates de tarisement : {e}")
            dry_dates = {}

        return dict(sorted(dry_dates.items(), key=lambda item : item[1]))

    def get_all_calving_preparation_date(self) -> dict[int, date]:
        """Retrieves and sorts the calving preparation dates for all cows with valid reproduction records.

        This function collects the 'calving_preparation' date for each cow and returns a dictionary sorted by date.

        Returns:
            dict[int, date]: A dictionary mapping cow IDs to their calving preparation dates, sorted by date.
        """

        calving_preparation_dates : dict[int, date] = {
            cow_id: parse_date(reproduction["calving_preparation"])
            for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=self.id).items()
            if not reproduction["calving_preparation_status"] and reproduction["calving_preparation"] is not None
        }

        return dict(sorted(calving_preparation_dates.items(), key=lambda item: item[1]))

    def get_all_calving_date(self) -> dict[int, date]:
        """Retrieves and sorts the calving dates for all cows with valid reproduction records.

        This function collects the 'calving_date' for each cow and returns a dictionary sorted by date.

        Returns:
            dict[int, date]: A dictionary mapping cow IDs to their calving dates, sorted by date.
        """

        calving_dates : dict[int, date] = {
            cow_id: parse_date(reproduction["calving_date"])
            for cow_id, reproduction in CowUtils.get_valid_reproduction(user_id=self.id).items()
            if reproduction["calving_date"] is not None
        }

        return dict(sorted(calving_dates.items(), key=lambda item: item[1]))
